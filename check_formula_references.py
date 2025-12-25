#!/usr/bin/env python3
"""
Check what the Financial Statements formulas actually reference.
"""

from openpyxl import load_workbook
from pathlib import Path

def check_formula_references(output_path):
    """Check what cells the Financial Statements formulas reference."""
    
    print(f"Loading output: {output_path}")
    wb = load_workbook(output_path, data_only=False)
    
    if 'Financial Statements' not in wb.sheetnames:
        print("Financial Statements sheet not found")
        return
    
    fs_sheet = wb['Financial Statements']
    case_data_sheet = wb['Case Data']
    
    print("\n" + "="*80)
    print("ANALYZING FORMULA REFERENCES IN FINANCIAL STATEMENTS SHEET")
    print("="*80)
    
    # Check the cells we saw in the comparison (B5, B15, B20)
    check_cells = ['B5', 'B15', 'B20', 'B83', 'B87', 'B90']
    
    print("\n--- Formula Chain Analysis ---")
    for cell_ref in check_cells:
        cell = fs_sheet[cell_ref]
        print(f"\n{cell_ref}:")
        print(f"  Data Type: {cell.data_type}")
        if cell.data_type == 'f':
            print(f"  Formula: {cell.value}")
            # Try to trace what it references
            formula = str(cell.value)
            if 'Case_Data' in formula:
                print(f"    -> References Case Data sheet")
            elif '!' in formula:
                # References another sheet
                print(f"    -> References another sheet")
            else:
                # Might reference same sheet
                print(f"    -> May reference same sheet (Financial Statements)")
        else:
            print(f"  Value: {cell.value}")
    
    # Check if B83, B87, B90 reference Case Data
    print("\n--- Checking Referenced Cells (B83, B87, B90) ---")
    for cell_ref in ['B83', 'B87', 'B90']:
        cell = fs_sheet[cell_ref]
        print(f"\n{cell_ref}:")
        print(f"  Data Type: {cell.data_type}")
        if cell.data_type == 'f':
            print(f"  Formula: {cell.value}")
            formula = str(cell.value)
            if 'Case_Data' in formula:
                print(f"    -> REFERENCES Case Data!")
                # Extract the Case Data reference
                if 'Case_Data!' in formula:
                    import re
                    matches = re.findall(r'Case_Data![A-Z]+\d+', formula)
                    for match in matches:
                        case_data_ref = match.replace('Case_Data!', '')
                        case_data_value = case_data_sheet[case_data_ref].value
                        print(f"      {match} = {case_data_value}")
        else:
            print(f"  Value: {cell.value}")
    
    # Check first few rows for company name or header
    print("\n--- Checking First 5 Rows for Headers ---")
    for row in range(1, 6):
        row_values = []
        for col in range(1, 6):  # A-E
            cell = fs_sheet.cell(row=row, column=col)
            if cell.value:
                row_values.append(f"{chr(64+col)}{row}={cell.value}")
        if row_values:
            print(f"Row {row}: {', '.join(row_values)}")
    
    # Check what Case Data B13 actually contains
    print("\n--- Case Data Values ---")
    print(f"B10 (Company): {case_data_sheet['B10'].value}")
    print(f"B13 (Sales/Revenue): {case_data_sheet['B13'].value}")
    print(f"B14 (COGS): {case_data_sheet['B14'].value}")
    print(f"B15 (R&D): {case_data_sheet['B15'].value}")
    
    # Check if values are in millions or raw dollars
    print("\n--- Value Magnitude Check ---")
    template_wb = load_workbook('templates/financial_analysis_template.xlsx', data_only=False)
    template_case_data = template_wb['Case Data']
    print(f"\nTemplate B13: {template_case_data['B13'].value}")
    print(f"Output B13: {case_data_sheet['B13'].value}")
    if case_data_sheet['B13'].value and template_case_data['B13'].value:
        ratio = case_data_sheet['B13'].value / template_case_data['B13'].value
        print(f"Ratio: {ratio:.2f}x")
        if ratio > 1000:
            print(f"  -> Output value is {ratio/1000:.0f} thousand times larger (likely NOT converted to millions)")
    
    wb.close()
    template_wb.close()

if __name__ == '__main__':
    script_dir = Path(__file__).parent.absolute()
    output_path = script_dir / 'Tool_Output' / 'Zoom-Communications-Inc-FY-2025-Financials.xlsx'
    check_formula_references(str(output_path))











