#!/usr/bin/env python3
"""
SEC Financials Tool
Fetches company financial data from SEC EDGAR API and creates Excel files
similar to the Tesla FY 2024 Financials format.
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from edgar import Company, set_identity
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
import json

# Suppress edgar library verbose logging
logging.getLogger("edgar").setLevel(logging.WARNING)


def get_company_financials(ticker: str, year: int = None, form_type: str = "10-K"):
    """
    Fetch financial statements for a given company ticker.
    
    Args:
        ticker: Company stock ticker symbol (e.g., 'TSLA')
        year: Fiscal year (optional, defaults to latest)
        form_type: Form type to fetch (default: '10-K' for annual reports)
    
    Returns:
        Dictionary containing income statement, balance sheet, and cash flow statement
    """
    try:
        company = Company(ticker)
        # Get filings, excluding amendments to ensure full XBRL data
        filings = company.get_filings(form=form_type, amendments=False)
        
        if filings.empty:
            raise ValueError(f"No {form_type} filings found for {ticker}")
        
        # Get the latest filing or specific year if provided
        if year:
            # Filter by year if specified
            filing = filings[filings.filing_date.dt.year == year]
            if filing.empty:
                raise ValueError(f"No {form_type} filing found for {ticker} in year {year}")
            filing = filing.iloc[0]
        else:
            filing = filings.latest()
        
        filing_date = filing.filing_date
        if hasattr(filing_date, 'date'):
            filing_date = filing_date.date()
        print(f"Fetching financials from {filing_date} filing...")
        
        # Get the filing object and extract financials
        filing_obj = filing.obj()
        financials = filing_obj.financials
        
        # Extract the three main financial statements
        income_statement = financials.income_statement()
        balance_sheet = financials.balance_sheet()
        cash_flow_statement = financials.cashflow_statement()
        
        return {
            'income_statement': income_statement,
            'balance_sheet': balance_sheet,
            'cash_flow_statement': cash_flow_statement,
            'filing_date': filing.filing_date,
            'company_name': company.name if hasattr(company, 'name') else ticker,
            'ticker': ticker
        }
    
    except Exception as e:
        raise Exception(f"Error fetching financials for {ticker}: {str(e)}")


def format_financial_dataframe(stmt) -> pd.DataFrame:
    """
    Format financial statement to DataFrame and match Excel structure.
    Handles both Statement objects and DataFrames.
    Extracts the most relevant columns (label and date values).
    """
    if stmt is None:
        return pd.DataFrame()
    
    # Convert Statement object to DataFrame if needed
    if not isinstance(stmt, pd.DataFrame):
        if hasattr(stmt, 'to_dataframe'):
            df = stmt.to_dataframe()
        else:
            return pd.DataFrame()
    else:
        df = stmt
    
    if df.empty:
        return pd.DataFrame()
    
    # Filter out abstract/metadata rows (rows where abstract is True or label is missing)
    if 'abstract' in df.columns:
        df = df[df['abstract'] != True]
    if 'label' in df.columns:
        df = df[df['label'].notna()]
    
    # Extract relevant columns: label and date columns
    relevant_cols = []
    if 'label' in df.columns:
        relevant_cols.append('label')
    
    # Find date columns (columns that look like dates: YYYY-MM-DD format)
    import re
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
    for col in df.columns:
        if date_pattern.match(str(col)) or (isinstance(col, str) and col.startswith('20')):
            relevant_cols.append(col)
    
    # If we found relevant columns, use them; otherwise use all columns
    if relevant_cols:
        df = df[relevant_cols].copy()
        # Rename label to Item for consistency
        if 'label' in df.columns:
            df = df.rename(columns={'label': 'Item'})
    else:
        # Fallback: reset index and use all columns
        if df.index.nlevels > 1:
            df.index = df.index.map(lambda x: ' - '.join(str(i) for i in x) if isinstance(x, tuple) else str(x))
        df = df.reset_index()
        if len(df.columns) > 0:
            df.columns.values[0] = 'Item'
    
    return df


def format_number_to_millions(value):
    """Convert a number to millions, handling None and NaN values."""
    if pd.isna(value) or value is None:
        return None
    try:
        num_value = float(value)
        if num_value == 0:
            return 0
        return num_value / 1_000_000
    except (ValueError, TypeError):
        return None


def format_date_for_header(date_str):
    """Format a date string (YYYY-MM-DD) to 'Year ended Month DD, YYYY' format."""
    try:
        from datetime import datetime
        date_obj = datetime.strptime(str(date_str), '%Y-%m-%d')
        return date_obj.strftime('Year ended %B %d, %Y')
    except:
        # Fallback: try to extract year from string
        if isinstance(date_str, str) and len(date_str) >= 4:
            year = date_str[:4]
            return f'Year ended {year}'
        return str(date_str)


def get_default_template_path():
    """
    Get the default path to the financial analysis template file.
    
    Returns:
        str: Path to the default template file (templates/financial_analysis_template.xlsx)
    """
    # Get the directory where this script is located
    script_dir = Path(__file__).parent.absolute()
    return str(script_dir / 'templates' / 'financial_analysis_template.xlsx')


def load_template(template_path: str):
    """
    Load an Excel template file using openpyxl.
    
    Args:
        template_path: Path to the template Excel file
        
    Returns:
        openpyxl.Workbook: The loaded workbook object
        
    Raises:
        FileNotFoundError: If the template file does not exist at the specified path
        Exception: If there's an error loading the workbook
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template file not found at {template_path}. "
            "Please ensure the template exists or provide a custom template_path."
        )
    
    try:
        workbook = load_workbook(template_path)
        return workbook
    except Exception as e:
        raise Exception(f"Error loading template file {template_path}: {str(e)}")


def validate_template_structure(workbook):
    """
    Validate that the template workbook contains all required sheets and structure.
    
    Args:
        workbook: openpyxl.Workbook object to validate
        
    Raises:
        ValueError: If required sheets are missing
        
    Note:
        Case Data sheet is created dynamically, so it is not required in the template.
    """
    # Required sheet names (Case Data will be created dynamically)
    required_sheets = [
        'Financial Statements',
        'Credit Analysis',
        'Forecasting Assumptions',
        'Valuation Parameters',
        'Residual Income Valuations',
        'DCF Valuations',
        'EPS Forecaster'
    ]
    
    # Get existing sheet names
    existing_sheets = workbook.sheetnames
    
    # Check for missing sheets
    missing_sheets = [sheet for sheet in required_sheets if sheet not in existing_sheets]
    
    if missing_sheets:
        raise ValueError(
            f"Template is missing required sheets: {', '.join(missing_sheets)}. "
            "Please ensure the template contains all required sheets."
        )
    
    # Note: Case Data sheet is created dynamically, so we don't validate it here


def get_default_config():
    """
    Get default configuration values for analysis parameters.
    
    Returns:
        dict: Dictionary containing default configuration values
    """
    return {
        'valuation_parameters': {
            'wacc': 0.10,
            'terminal_growth_rate': 0.03,
            'risk_free_rate': 0.025
        },
        'forecasting_assumptions': {
            'revenue_growth_rate': 0.15,
            'operating_margin': 0.20
        },
        'cell_mappings': {}
    }


def load_analysis_config(config_path: str = None):
    """
    Load analysis configuration from a JSON file, or return defaults if file doesn't exist.
    
    Args:
        config_path: Optional path to the configuration file. If None, looks for
                    'analysis_config.json' in the project root.
    
    Returns:
        dict: Configuration dictionary, either from file or defaults
        
    Raises:
        ValueError: If the config file exists but contains invalid JSON or cannot be read
    """
    # Determine config file path
    if config_path is None:
        script_dir = Path(__file__).parent.absolute()
        config_path = str(script_dir / 'analysis_config.json')
    
    # Check if config file exists
    if not os.path.exists(config_path):
        # Return default config if file doesn't exist
        return get_default_config()
    
    # Try to read and parse the config file
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Merge with defaults to ensure all required keys exist
        default_config = get_default_config()
        # Merge default_config into config (config values take precedence)
        merged_config = default_config.copy()
        merged_config.update(config)
        
        # Ensure nested dictionaries are also merged
        if 'valuation_parameters' in config:
            merged_config['valuation_parameters'] = {
                **default_config.get('valuation_parameters', {}),
                **config['valuation_parameters']
            }
        if 'forecasting_assumptions' in config:
            merged_config['forecasting_assumptions'] = {
                **default_config.get('forecasting_assumptions', {}),
                **config['forecasting_assumptions']
            }
        if 'cell_mappings' in config:
            merged_config['cell_mappings'] = config['cell_mappings']
        
        return merged_config
        
    except json.JSONDecodeError as e:
        raise ValueError(
            f"Error reading configuration file: Invalid JSON syntax in {config_path}. "
            f"JSON parsing error: {str(e)}. Please check the file format and try again."
        )
    except PermissionError as e:
        raise ValueError(
            f"Error reading configuration file: Permission denied for {config_path}. "
            f"Please check file permissions and try again."
        )
    except Exception as e:
        raise ValueError(
            f"Error reading configuration file: {str(e)}. "
            "Please check the file format and try again."
        )


def _format_case_data_sheet(case_data_sheet):
    """
    Apply formatting to the Case Data sheet.
    
    Args:
        case_data_sheet: openpyxl.Worksheet object for Case Data sheet
    """
    # Format header row (row 9)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, 7):  # Columns A-F
        cell = case_data_sheet.cell(row=9, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format label column (Column A)
    for row in range(10, 44):
        cell = case_data_sheet.cell(row=row, column=1)
        if cell.value:  # Only format if cell has a value
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
    
    # Format data columns (Columns B-F) - right align numbers
    for row in range(10, 44):
        for col in range(2, 7):
            cell = case_data_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='right')
            # Apply number format if value is numeric
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    
    # Set column widths
    case_data_sheet.column_dimensions['A'].width = 35
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        case_data_sheet.column_dimensions[col_letter].width = 15


def create_case_data_sheet(workbook, financials_data: dict):
    """
    Create a new Case Data sheet with proper structure and populate with SEC financial data.
    
    Args:
        workbook: openpyxl.Workbook object
        financials_data: Dictionary from get_company_financials() containing financial data
    """
    # Remove existing Case Data sheet if it exists
    if 'Case Data' in workbook.sheetnames:
        workbook.remove(workbook['Case Data'])
    
    # Create new Case Data sheet
    case_data_sheet = workbook.create_sheet('Case Data', index=0)  # Insert at beginning
    
    # Define row labels (Column A)
    # Note: Row 10 labels are added separately (Company Name, Ticker, SIC in A, C, E)
    #       Row 11 and 12 have labels in Column A only
    row_labels = {
        11: ['Common Shares Outstanding'],  # A11 - label, B11 will have value
        12: ['Fiscal Year End'],  # A12 - label, B12 will have value
        13: ['Sales'],
        14: ['Cost of Goods Sold'],
        15: ['R&D'],
        16: ['SG&A'],
        17: ['Depreciation'],
        18: ['Interest Expense'],
        19: ['Non-Operating Income'],
        20: ['Income Tax'],
        21: ['Noncontrolling Interest'],
        22: ['Other Income'],
        23: ['Extraordinary Items'],
        24: ['Preferred Dividends'],
        25: ['Cash and Cash Equivalents'],
        26: ['Receivables'],
        27: ['Inventories'],
        28: ['Other Current Assets'],
        29: ['Property, Plant and Equipment'],
        30: ['Investments'],
        31: ['Intangible Assets'],
        32: ['Other Assets'],
        33: ['Current Debt'],
        34: ['Accounts Payable'],
        35: ['Income Taxes Payable'],
        36: ['Other Current Liabilities'],
        37: ['Long-Term Debt'],
        38: ['Other Liabilities'],
        39: ['Deferred Taxes'],
        40: ['Noncontrolling Interest'],
        41: ['Preferred Stock'],
        42: ['Paid in Capital'],
        43: ['Retained Earnings'],
    }
    
    # Populate row labels in Column A (and other columns for row 10)
    for row_num, labels in row_labels.items():
        for col_offset, label in enumerate(labels):
            if label:  # Only write non-empty labels
                case_data_sheet.cell(row=row_num, column=1 + col_offset, value=label)
    
    # Add labels for row 10 (Company Name, Ticker, SIC)
    case_data_sheet.cell(row=10, column=1, value='Company Name')  # A10
    case_data_sheet.cell(row=10, column=3, value='Ticker')  # C10
    case_data_sheet.cell(row=10, column=5, value='SIC')  # E10
    
    # Add column headers for years (Row 9)
    case_data_sheet.cell(row=9, column=1, value='Item')
    for col_idx, year_label in enumerate(['Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5'], start=2):
        case_data_sheet.cell(row=9, column=col_idx, value=year_label)
    
    # Populate metadata (rows 10-12)
    populate_case_data_metadata(case_data_sheet, financials_data)
    
    # Format financial statements to DataFrames
    income_df = format_financial_dataframe(financials_data.get('income_statement'))
    balance_df = format_financial_dataframe(financials_data.get('balance_sheet'))
    cash_flow_df = format_financial_dataframe(financials_data.get('cash_flow_statement'))
    
    # Map data to Case Data sheet
    if not income_df.empty:
        map_income_statement_to_case_data(income_df, workbook)
    
    if not balance_df.empty:
        map_balance_sheet_to_case_data(balance_df, workbook)
    
    if not cash_flow_df.empty:
        map_cash_flow_to_case_data(cash_flow_df, workbook)
    
    # Apply formatting
    _format_case_data_sheet(case_data_sheet)


def populate_case_data_metadata(case_data_sheet, financials_data: dict):
    """
    Populate metadata rows (10-12) in the Case Data sheet with company information.
    
    Args:
        case_data_sheet: openpyxl.Worksheet object for Case Data sheet
        financials_data: Dictionary containing financial data from get_company_financials()
                        Expected keys: 'company_name', 'filing_date', and optionally 'ticker'
    """
    company_name = financials_data.get('company_name', '')
    ticker = financials_data.get('ticker', '')
    
    # Row 10: Company Name, Ticker, SIC
    # Cell A10: Label "Company Name"
    # Cell B10: Company Name
    case_data_sheet['B10'] = company_name
    
    # Cell D10: Ticker (if available)
    if ticker:
        case_data_sheet['D10'] = ticker
    # Note: SIC (cell F10) is typically not available from SEC API, so we leave it empty
    
    # Row 11: Common Shares Outstanding
    # This is typically not directly available from the SEC API in the format we get
    # We'll leave it empty for now
    # If balance sheet has share information, it would need to be extracted separately
    
    # Row 12: Fiscal Year End dates (columns B-F for historical years)
    filing_date = financials_data.get('filing_date')
    
    if filing_date:
        # Extract year from filing_date
        if hasattr(filing_date, 'year'):
            fiscal_year = filing_date.year
        elif hasattr(filing_date, 'date'):
            fiscal_year = filing_date.date().year
        else:
            # Try to extract from string
            try:
                fiscal_year = int(str(filing_date)[:4])
            except (ValueError, TypeError):
                fiscal_year = None
        
        if fiscal_year:
            # Populate fiscal year end date in column B (most recent year)
            # Set to December 31 of the fiscal year (common fiscal year end date)
            try:
                fiscal_year_end = datetime(fiscal_year, 12, 31)
                case_data_sheet['B12'] = fiscal_year_end
            except (ValueError, TypeError):
                # If date creation fails, just set the year as string
                case_data_sheet['B12'] = f"{fiscal_year}-12-31"


def _find_matching_row(df, search_terms, case_sensitive=False):
    """
    Helper function to find a row in a DataFrame that matches one of the search terms.
    
    Args:
        df: pandas DataFrame with an 'Item' column
        search_terms: List of strings to search for (case-insensitive by default)
        case_sensitive: Whether to do case-sensitive matching
        
    Returns:
        pandas Series or None: The matching row, or None if not found
    """
    if df.empty or 'Item' not in df.columns:
        return None
    
    for term in search_terms:
        if case_sensitive:
            mask = df['Item'].str.contains(term, na=False, regex=False)
        else:
            mask = df['Item'].str.contains(term, na=False, regex=False, case=False)
        if mask.any():
            return df[mask].iloc[0]
    return None


def _get_value_from_row(row, column_index=0):
    """
    Helper function to extract a value from a DataFrame row.
    
    Args:
        row: pandas Series (a row from a DataFrame)
        column_index: Index of the numeric column to extract (0 = first date column, 1 = second, etc.)
    
    Returns:
        float or None: The numeric value, or None if not available
    """
    if row is None or row.empty:
        return None
    
    # Get all columns excluding 'Item'
    numeric_cols = [col for col in row.index if col != 'Item']
    
    # Filter to only numeric-like columns (those that can be converted to float)
    valid_numeric_cols = []
    for col in numeric_cols:
        try:
            val = row[col]
            if pd.notna(val):
                # Try to convert to float to verify it's numeric
                float(val)
                valid_numeric_cols.append(col)
        except (ValueError, TypeError):
            continue
    
    # Try to get the value from the specified column index
    if valid_numeric_cols:
        try:
            col_name = valid_numeric_cols[column_index] if column_index < len(valid_numeric_cols) else valid_numeric_cols[0]
            value = row[col_name]
            if pd.notna(value):
                return float(value)
        except (IndexError, ValueError, TypeError):
            pass
    
    return None


def map_income_statement_to_case_data(income_df, workbook):
    """
    Map Income Statement data to the Case Data sheet.
    
    Args:
        income_df: pandas DataFrame containing Income Statement data (from format_financial_dataframe)
        workbook: openpyxl.Workbook object containing the template
        
    Note:
        Values are converted to millions before writing to Case Data.
        Maps to rows 13-24 in Case Data sheet.
    """
    if income_df.empty or 'Case Data' not in workbook.sheetnames:
        return
    
    case_data_sheet = workbook['Case Data']
    
    # Mapping of row numbers to search terms for Income Statement items
    # Each list contains common variations of the label name
    income_statement_mapping = {
        13: ['Sales', 'Revenue', 'Net Sales', 'Total Revenue', 'Revenues'],
        14: ['Cost of Goods Sold', 'Cost of Revenue', 'Cost of Sales', 'COGS', 'Cost of products sold'],
        15: ['Research and Development', 'R&D', 'Research and development expense', 'Research & Development'],
        16: ['Selling, General and Administrative', 'SG&A', 'Selling general and administrative', 'Selling, general & administrative'],
        17: ['Depreciation', 'Amortization', 'Depreciation and Amortization', 'Depreciation & Amortization'],
        18: ['Interest Expense', 'Interest expense', 'Interest and debt expense'],
        19: ['Non-Operating Income', 'Nonoperating Income', 'Other Income', 'Non-operating income'],
        20: ['Income Tax', 'Income Taxes', 'Tax expense', 'Provision for income taxes'],
        21: ['Noncontrolling Interest', 'Non-controlling Interest', 'Noncontrolling interests'],
        22: ['Other Income', 'Other income (loss)', 'Other Income (Loss)'],
        23: ['Extraordinary Items', 'Discontinued Operations', 'Ext. Items', 'Extraordinary items'],
        24: ['Preferred Dividends', 'Preferred Stock Dividends', 'Preferred dividends']
    }
    
    # Process each row mapping
    for row_num, search_terms in income_statement_mapping.items():
        matching_row = _find_matching_row(income_df, search_terms)
        if matching_row is not None:
            # Get the value from the first date column (most recent year)
            value = _get_value_from_row(matching_row, column_index=0)
            
            if value is not None:
                # Convert to millions before writing to Case Data
                value_in_millions = format_number_to_millions(value)
                # Write to column B (most recent year)
                # Note: Some items like expenses should be negative, but SEC data typically already includes sign
                case_data_sheet.cell(row=row_num, column=2, value=value_in_millions)


def map_balance_sheet_to_case_data(balance_df, workbook):
    """
    Map Balance Sheet data to the Case Data sheet.
    
    Args:
        balance_df: pandas DataFrame containing Balance Sheet data (from format_financial_dataframe)
        workbook: openpyxl.Workbook object containing the template
        
    Note:
        Values are converted to millions before writing to Case Data.
        Maps to rows 25+ in Case Data sheet.
    """
    if balance_df.empty or 'Case Data' not in workbook.sheetnames:
        return
    
    case_data_sheet = workbook['Case Data']
    
    # Mapping of row numbers to search terms for Balance Sheet items
    # Starting from row 25 as per template structure
    balance_sheet_mapping = {
        25: ['Cash and Cash Equivalents', 'Cash and cash equivalents', 'Operating Cash', 'Cash'],
        26: ['Receivables', 'Accounts Receivable', 'Trade Receivables', 'Accounts receivable, net'],
        27: ['Inventories', 'Inventory', 'Inventories, net'],
        28: ['Other Current Assets', 'Other current assets', 'Prepaid Expenses'],
        29: ['Property, Plant and Equipment', 'PP&E', 'Property and equipment', 'Property, plant and equipment, net'],
        30: ['Investments', 'Investment', 'Available-for-sale securities', 'Short-term investments'],
        31: ['Intangible Assets', 'Intangibles', 'Goodwill', 'Goodwill and intangible assets'],
        32: ['Other Assets', 'Other non-current assets', 'Other assets, net'],
        33: ['Current Debt', 'Short-term Debt', 'Current portion of long-term debt', 'Debt, current'],
        34: ['Accounts Payable', 'Trade Payables', 'Accounts payable'],
        35: ['Income Taxes Payable', 'Taxes Payable', 'Income tax payable'],
        36: ['Other Current Liabilities', 'Other current liabilities', 'Accrued Liabilities'],
        37: ['Long-Term Debt', 'Long-term debt', 'Debt, non-current', 'Long term debt'],
        38: ['Other Liabilities', 'Other non-current liabilities', 'Other liabilities'],
        39: ['Deferred Taxes', 'Deferred Tax', 'Deferred income taxes', 'Deferred tax liabilities'],
        40: ['Noncontrolling Interest', 'Non-controlling Interest', 'Noncontrolling interests'],
        41: ['Preferred Stock', 'Preferred stock'],
        42: ['Paid in Capital', 'Common Stock', 'Share Capital', 'Paid-in capital'],
        43: ['Retained Earnings', 'Retained earnings', 'Accumulated deficit']
    }
    
    # Process each row mapping
    for row_num, search_terms in balance_sheet_mapping.items():
        matching_row = _find_matching_row(balance_df, search_terms)
        if matching_row is not None:
            value = _get_value_from_row(matching_row, column_index=0)
            if value is not None:
                # Convert to millions before writing to Case Data
                value_in_millions = format_number_to_millions(value)
                # Write to column B (most recent year)
                case_data_sheet.cell(row=row_num, column=2, value=value_in_millions)


def map_cash_flow_to_case_data(cash_flow_df, workbook):
    """
    Map Cash Flow Statement data to the Case Data sheet.
    
    Args:
        cash_flow_df: pandas DataFrame containing Cash Flow Statement data (from format_financial_dataframe)
        workbook: openpyxl.Workbook object containing the template
        
    Note:
        Cash Flow items may not have dedicated rows in the Case Data sheet structure.
        The template typically calculates cash flows from Income Statement and Balance Sheet changes.
        This function is provided for completeness and can be extended if specific cash flow rows are needed.
    """
    # Cash Flow Statement items typically don't have dedicated rows in Case Data
    # The template calculates cash flows from changes in Balance Sheet items
    # If specific cash flow rows are needed in the future, they can be added here
    # For now, this function is a placeholder that does nothing
    pass


def populate_financial_statements_raw_data(workbook, financials_data: dict):
    """
    Populate the Financial Statements sheet Raw Data Inputs section (rows 82-94)
    with data from Case Data sheet.
    
    The Financial Statements sheet has a "Raw Data Inputs" section (rows 82-94) that
    contains hardcoded template values. The display formulas in rows 1-79 reference
    these cells (e.g., B4 references B82 for company name, B5 references B83 for shares).
    This function replaces the template values with the new company's data from Case Data.
    
    Args:
        workbook: openpyxl.Workbook object containing the template
        financials_data: Dictionary from get_company_financials() containing:
                        - 'company_name': str
                        - 'ticker': str
    """
    if 'Financial Statements' not in workbook.sheetnames or 'Case Data' not in workbook.sheetnames:
        return
    
    fs_sheet = workbook['Financial Statements']
    case_data = workbook['Case Data']
    
    # B82: Company Name and Ticker (combine from Case Data B10 and D10)
    company_name = case_data['B10'].value or ''
    ticker = case_data['D10'].value or ''
    # Format: "COMPANY NAME TICKER" (e.g., "Zoom Communications, Inc. ZM")
    company_name_ticker = f"{company_name} {ticker}".strip()
    fs_sheet['B82'] = company_name_ticker
    
    # B83: Common Shares Outstanding (from Case Data B11)
    fs_sheet['B83'] = case_data['B11'].value
    
    # B84: Fiscal Year End (from Case Data B12)
    fs_sheet['B84'] = case_data['B12'].value
    
    # B85-B94: Income Statement items (from Case Data B13-B22)
    # These values should already be in millions from Case Data population
    fs_sheet['B85'] = case_data['B13'].value  # Sales (Net)
    fs_sheet['B86'] = case_data['B14'].value  # Cost of Goods Sold
    fs_sheet['B87'] = case_data['B15'].value  # R&D Expense
    fs_sheet['B88'] = case_data['B16'].value  # SG&A Expense
    fs_sheet['B89'] = case_data['B17'].value  # Depreciation & Amortization
    fs_sheet['B90'] = case_data['B18'].value  # Interest Expense
    fs_sheet['B91'] = case_data['B19'].value  # Non-Operating Income (Loss)
    fs_sheet['B92'] = case_data['B20'].value  # Income Taxes
    fs_sheet['B93'] = case_data['B21'].value  # Noncontrolling Interest in Earnings
    fs_sheet['B94'] = case_data['B22'].value  # Other Income (Loss)


def populate_case_data_sheet(workbook, financials_data: dict):
    """
    Create and populate the Case Data sheet with SEC financial data.
    
    This function creates a new Case Data sheet from scratch and populates it with:
    1. Metadata (company name, ticker, dates, shares)
    2. Income Statement data
    3. Balance Sheet data
    4. Cash Flow Statement data (if needed)
    
    Args:
        workbook: openpyxl.Workbook object containing the template
        financials_data: Dictionary from get_company_financials() containing:
                        - 'income_statement': Statement object or DataFrame
                        - 'balance_sheet': Statement object or DataFrame
                        - 'cash_flow_statement': Statement object or DataFrame
                        - 'company_name': str
                        - 'ticker': str
                        - 'filing_date': date or datetime
    
    Note:
        This function creates the Case Data sheet from scratch, so it should be called
        even if the template doesn't have a Case Data sheet.
    """
    # Create Case Data sheet from scratch
    create_case_data_sheet(workbook, financials_data)


def populate_config_values(workbook, config: dict):
    """
    Populate configuration parameter values into specified cells in analysis sheets.
    
    Args:
        workbook: openpyxl.Workbook object containing the template
        config: Configuration dictionary with structure:
                {
                    'valuation_parameters': {...},
                    'forecasting_assumptions': {...},
                    'cell_mappings': {
                        'Sheet Name': {
                            'B2': 'parameter_key',
                            'C3': 'another_parameter_key'
                        }
                    }
                }
    
    Note:
        Only writes values to cells, preserving existing formatting.
        Does not modify formulas - only updates cell values.
    """
    if 'cell_mappings' not in config:
        # No cell mappings defined, nothing to populate
        return
    
    cell_mappings = config['cell_mappings']
    
    # Create a flattened parameter lookup dictionary
    # This combines valuation_parameters and forecasting_assumptions
    param_lookup = {}
    if 'valuation_parameters' in config:
        param_lookup.update(config['valuation_parameters'])
    if 'forecasting_assumptions' in config:
        param_lookup.update(config['forecasting_assumptions'])
    
    # Process each sheet's cell mappings
    for sheet_name, cell_map in cell_mappings.items():
        if sheet_name not in workbook.sheetnames:
            # Skip silently or log warning - sheet doesn't exist
            continue
        
        worksheet = workbook[sheet_name]
        
        # Process each cell mapping
        for cell_ref, param_key in cell_map.items():
            # Look up the parameter value
            if param_key in param_lookup:
                value = param_lookup[param_key]
                
                # Write the value to the cell (preserves formatting)
                try:
                    worksheet[cell_ref] = value
                except Exception as e:
                    # Skip invalid cell references with a warning (could log this in production)
                    continue


def format_sheet_with_headers(writer, sheet_name, df, company_name, report_type, fiscal_year):
    """
    Format a financial statement sheet with proper headers, spacing, and number formatting.
    """
    if df.empty:
        return
    
    # Create a copy to avoid modifying original
    formatted_df = df.copy()
    
    # Convert numeric columns (date columns) to millions
    numeric_cols = []
    date_headers = []
    for col in formatted_df.columns:
        if col != 'Item':
            # Check if column is numeric or looks like a date column (YYYY-MM-DD format)
            is_numeric = pd.api.types.is_numeric_dtype(formatted_df[col])
            is_date_col = isinstance(col, str) and len(str(col)) == 10 and str(col)[4] == '-' and str(col)[7] == '-'
            
            if is_numeric or is_date_col:
                numeric_cols.append(col)
                # Format date header
                date_headers.append(format_date_for_header(col) + ' (In millions)')
                # Convert values to millions
                formatted_df[col] = formatted_df[col].apply(format_number_to_millions)
    
    # Write to Excel starting at row 4 (0-indexed, so row 4 = Excel row 4)
    # Row 1: Company name
    # Row 2: Report type
    # Row 3: Empty (spacing)
    # Row 4: Column headers
    # Row 5+: Data
    start_row = 4
    formatted_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, header=False)
    
    # Get the workbook and worksheet for formatting
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Add company name header (row 1, Excel row 1)
    worksheet.cell(row=1, column=1, value=company_name.upper())
    worksheet.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    # Add report type header (row 2, Excel row 2)
    report_header = f"Annual Report - FY {fiscal_year} - {sheet_name}"
    worksheet.cell(row=2, column=1, value=report_header)
    worksheet.cell(row=2, column=1).font = Font(bold=True, size=11)
    
    # Row 3 is empty (spacing row) - already handled by startrow
    
    # Add column headers (row 4, Excel row 4)
    worksheet.cell(row=start_row, column=1, value='')  # Empty cell for Item column header
    for idx, header in enumerate(date_headers, start=2):
        cell = worksheet.cell(row=start_row, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
    
    # Format the data rows (starting at row 5)
    for row_idx in range(start_row + 1, start_row + 1 + len(formatted_df)):
        # Format Item column (column 1)
        item_cell = worksheet.cell(row=row_idx, column=1)
        item_cell.alignment = Alignment(horizontal='left')
        
        # Format numeric columns (right-aligned, comma formatting)
        for col_idx, col_name in enumerate(numeric_cols, start=2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None and pd.notna(cell.value):
                try:
                    # Check if value is numeric
                    num_value = float(cell.value)
                    # Format as number with commas, no decimals for whole numbers
                    if abs(num_value) >= 1:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    # If not numeric, leave as is
                    pass
            cell.alignment = Alignment(horizontal='right')
    
    # Add spacing rows for grouping (identify rows that should have spacing)
    # Look for rows where Item column contains a colon (indicating a section header)
    current_row = start_row + 1
    rows_to_insert = []
    for idx, row in formatted_df.iterrows():
        item_value = str(row['Item']) if pd.notna(row['Item']) else ''
        # If item ends with colon, add spacing after
        if item_value.strip().endswith(':'):
            rows_to_insert.append(current_row)
        current_row += 1
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 50  # Item column
    for col_idx in range(2, len(numeric_cols) + 2):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = 25


def create_excel_file(ticker: str, output_path: str = None, year: int = None, 
                      form_type: str = "10-K", user_email: str = None,
                      template_path: str = None, config_path: str = None):
    """
    Create an Excel file with company financial statements.
    
    If a template file is found, creates a comprehensive financial analysis workbook
    with analysis sheets. Otherwise, creates a simple workbook with just the three
    basic financial statements (backward compatible behavior).
    
    Args:
        ticker: Company stock ticker symbol
        output_path: Path for output Excel file (optional)
        year: Fiscal year (optional)
        form_type: Form type to fetch (default: '10-K')
        user_email: Email for SEC API identification (optional, will prompt if not provided)
        template_path: Optional path to Excel template file. If None, uses default template.
        config_path: Optional path to JSON configuration file. If None, uses default or built-in defaults.
    """
    # Set identity for SEC API (required)
    if user_email:
        set_identity(user_email)
    else:
        # Try to get from environment or use a default
        import os
        email = os.getenv('SEC_API_EMAIL', 'user@example.com')
        set_identity(email)
        print(f"Using email: {email} (set SEC_API_EMAIL env var or use --email to customize)")
    
    # Fetch financial data
    print(f"Fetching financial data for {ticker}...")
    financials_data = get_company_financials(ticker, year, form_type)
    
    # Format the data
    income_df = format_financial_dataframe(financials_data['income_statement'])
    balance_df = format_financial_dataframe(financials_data['balance_sheet'])
    cash_flow_df = format_financial_dataframe(financials_data['cash_flow_statement'])
    
    # Generate output filename if not provided
    if output_path is None:
        company_name = financials_data['company_name']
        # Clean company name for filename (remove commas, periods, etc.)
        company_name = company_name.replace(',', '').replace('.', '').replace(' ', '-')
        filing_date = financials_data['filing_date']
        if filing_date:
            if hasattr(filing_date, 'year'):
                filing_year = filing_date.year
            elif hasattr(filing_date, 'date'):
                filing_year = filing_date.date().year
            else:
                filing_year = year or datetime.now().year
        else:
            filing_year = year or datetime.now().year
        
        # Default to Tool_Output folder
        script_dir = Path(__file__).parent.absolute()
        output_dir = script_dir / 'Tool_Output'
        output_dir.mkdir(exist_ok=True)  # Create directory if it doesn't exist
        output_path = str(output_dir / f"{company_name}-FY-{filing_year}-Financials.xlsx")
    
    # Determine fiscal year for headers
    filing_date = financials_data['filing_date']
    if filing_date:
        if hasattr(filing_date, 'year'):
            fiscal_year = filing_date.year
        elif hasattr(filing_date, 'date'):
            fiscal_year = filing_date.date().year
        else:
            fiscal_year = year or datetime.now().year
    else:
        fiscal_year = year or datetime.now().year
    
    # Determine template path
    if template_path is None:
        template_path = get_default_template_path()
    
    # Check if template exists and use template-based workflow if available
    use_template = os.path.exists(template_path)
    
    if use_template:
        # Template-based workflow
        print(f"Using template-based workflow with template: {template_path}")
        print(f"Creating Excel file: {output_path}")
        
        try:
            # Load template
            workbook = load_template(template_path)
            
            # Validate template structure (Case Data will be created dynamically)
            validate_template_structure(workbook)
            
            # Create and populate Case Data sheet with SEC data
            populate_case_data_sheet(workbook, financials_data)
            
            # Populate Financial Statements Raw Data Inputs section (rows 82-94)
            # This replaces Dollar Tree template values with the new company's data
            populate_financial_statements_raw_data(workbook, financials_data)
            
            # Load configuration
            config = load_analysis_config(config_path)
            
            # Populate config values into analysis sheets
            populate_config_values(workbook, config)
            
            # Save the workbook
            workbook.save(output_path)
            
            print(f"✓ Successfully created {output_path} (template-based)")
            print(f"  Company: {financials_data['company_name']}")
            filing_date = financials_data['filing_date']
            if filing_date and hasattr(filing_date, 'date'):
                filing_date = filing_date.date()
            print(f"  Filing Date: {filing_date if filing_date else 'N/A'}")
            
            return output_path
            
        except Exception as e:
            # If template workflow fails, fall back to standard workflow
            print(f"Warning: Template-based workflow failed ({str(e)}), falling back to standard workflow")
            use_template = False
    
    if not use_template:
        # Standard workflow (backward compatible)
        print(f"Creating Excel file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if not income_df.empty:
                format_sheet_with_headers(
                    writer, 'Income Statement', income_df,
                    financials_data['company_name'], 'Income Statement', fiscal_year
                )
            
            if not balance_df.empty:
                format_sheet_with_headers(
                    writer, 'Balance Sheet', balance_df,
                    financials_data['company_name'], 'Balance Sheet', fiscal_year
                )
            
            if not cash_flow_df.empty:
                format_sheet_with_headers(
                    writer, 'Cash Flow Statement', cash_flow_df,
                    financials_data['company_name'], 'Cash Flow Statement', fiscal_year
                )
        
        print(f"✓ Successfully created {output_path}")
        print(f"  Company: {financials_data['company_name']}")
        filing_date = financials_data['filing_date']
        if filing_date and hasattr(filing_date, 'date'):
            filing_date = filing_date.date()
        print(f"  Filing Date: {filing_date if filing_date else 'N/A'}")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Fetch company financial data from SEC EDGAR API and create Excel files',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sec_financials_tool.py TSLA
  python sec_financials_tool.py AAPL --year 2023
  python sec_financials_tool.py MSFT --output Microsoft-Financials.xlsx
  python sec_financials_tool.py GOOGL --email your.email@example.com
        """
    )
    
    parser.add_argument('ticker', help='Company stock ticker symbol (e.g., TSLA, AAPL)')
    parser.add_argument('--year', type=int, help='Fiscal year (default: latest available)')
    parser.add_argument('--output', '-o', help='Output Excel file path (default: auto-generated)')
    parser.add_argument('--form', default='10-K', help='Form type (default: 10-K)')
    parser.add_argument('--email', help='Email for SEC API identification (or set SEC_API_EMAIL env var)')
    
    args = parser.parse_args()
    
    try:
        create_excel_file(
            ticker=args.ticker.upper(),
            output_path=args.output,
            year=args.year,
            form_type=args.form,
            user_email=args.email
        )
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()

