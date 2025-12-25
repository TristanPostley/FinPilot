#!/usr/bin/env python3
"""
Check what B82 in Financial Statements references.
"""

from openpyxl import load_workbook
from pathlib import Path

def check_b82_reference(output_path):
    """Check what B82 references."""
    
    wb = load_workbook(output_path, data_only=False)
    fs_sheet = wb['Financial Statements']
    case_data_sheet = wb['Case Data']
    
    print("Checking B82 and surrounding cells in Financial Statements sheet:")
    print("\n" + "="*80)
    
    # Check B82 and nearby cells
    for cell_ref in ['B80', 'B81', 'B82', 'B83', 'B84', 'B85', 'B86', 'B87', 'B88', 'B89', 'B90']:
        cell = fs_sheet[cell_ref]
        print(f"\n{cell_ref}:")
        print(f"  Data Type: {cell.data_type}")
        if cell.data_type == 'f':
            print(f"  Formula: {cell.value}")
            formula = str(cell.value)
            if 'Case_Data' in formula:
                print(f"    -> REFERENCES Case Data!")
                # Extract Case Data references
                import re
                matches = re.findall(r'Case_Data![A-Z]+\d+', formula)
                for match in matches:
                    case_data_ref = match.replace('Case_Data!', '')
                    try:
                        case_data_value = case_data_sheet[case_data_ref].value
                        print(f"      {match} = {case_data_value}")
                    except:
                        print(f"      {match} = (error reading cell)")
        else:
            print(f"  Value: {cell.value}")
    
    # Check row labels around row 82
    print("\n--- Row Labels around Row 82 ---")
    for row in range(80, 95):
        label_cell = fs_sheet.cell(row=row, column=1)  # Column A
        if label_cell.value:
            print(f"Row {row} (A{row}): {label_cell.value}")
    
    wb.close()

if __name__ == '__main__':
    script_dir = Path(__file__).parent.absolute()
    output_path = script_dir / 'Tool_Output' / 'Zoom-Communications-Inc-FY-2025-Financials.xlsx'
    check_b82_reference(str(output_path))











