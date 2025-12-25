#!/usr/bin/env python3
"""
Check the mapping between Case Data rows and Financial Statements Raw Data Input rows.
"""

from openpyxl import load_workbook
from pathlib import Path

def check_row_mapping(template_path):
    """Check what rows in Case Data correspond to Financial Statements Raw Data Input rows."""
    
    wb = load_workbook(template_path, data_only=False)
    case_data = wb['Case Data']
    fs_sheet = wb['Financial Statements']
    
    print("="*80)
    print("ROW MAPPING: Case Data -> Financial Statements Raw Data Inputs")
    print("="*80)
    
    # Financial Statements Raw Data Input section (rows 82-94)
    fs_raw_data_rows = {
        82: "Company Name and Ticker",
        83: "Common Shares Outstanding",
        84: "Fiscal Year End",
        85: "Sales (Net)",
        86: "Cost of Goods Sold",
        87: "R&D Expense",
        88: "SG&A Expense",
        89: "Depreciation & Amortization",
        90: "Interest Expense",
        91: "Non-Operating Income (Loss)",
        92: "Income Taxes",
        93: "Noncontrolling Interest in Earnings",
        94: "Other Income (Loss)",
    }
    
    # Case Data income statement rows (13-24)
    case_data_income_rows = {
        13: "Sales",
        14: "Cost of Goods Sold",
        15: "R&D",
        16: "SG&A",
        17: "Depreciation",
        18: "Interest Expense",
        19: "Non-Operating Income",
        20: "Income Tax",
        21: "Noncontrolling Interest",
        22: "Other Income",
        23: "Extraordinary Items",
        24: "Preferred Dividends",
    }
    
    print("\n--- Financial Statements Raw Data Input Rows (82-94) ---")
    for fs_row, label in fs_raw_data_rows.items():
        cell_b = fs_sheet.cell(row=fs_row, column=2)  # Column B
        cell_a = fs_sheet.cell(row=fs_row, column=1)  # Column A (label)
        
        print(f"\nRow {fs_row} ({label}):")
        print(f"  Column A: {cell_a.value}")
        print(f"  Column B: {cell_b.value} (Type: {cell_b.data_type})")
        
        # Try to find corresponding Case Data row
        if fs_row == 82:
            print(f"    -> Should map to Case Data B10 (Company Name) and D10 (Ticker)")
        elif fs_row == 83:
            print(f"    -> Should map to Case Data B11 (Shares Outstanding)")
        elif fs_row == 84:
            print(f"    -> Should map to Case Data B12 (Fiscal Year End)")
        elif fs_row == 85:
            print(f"    -> Should map to Case Data B13 (Sales)")
        elif fs_row == 86:
            print(f"    -> Should map to Case Data B14 (COGS)")
        elif fs_row == 87:
            print(f"    -> Should map to Case Data B15 (R&D)")
        elif fs_row == 88:
            print(f"    -> Should map to Case Data B16 (SG&A)")
        elif fs_row == 89:
            print(f"    -> Should map to Case Data B17 (Depreciation)")
        elif fs_row == 90:
            print(f"    -> Should map to Case Data B18 (Interest Expense)")
        elif fs_row == 91:
            print(f"    -> Should map to Case Data B19 (Non-Operating Income)")
        elif fs_row == 92:
            print(f"    -> Should map to Case Data B20 (Income Taxes)")
        elif fs_row == 93:
            print(f"    -> Should map to Case Data B21 (Noncontrolling Interest)")
        elif fs_row == 94:
            print(f"    -> Should map to Case Data B22 (Other Income)")
    
    print("\n" + "="*80)
    print("CASE DATA ROWS (13-24) - Income Statement")
    print("="*80)
    for cd_row, label in case_data_income_rows.items():
        cell_b = case_data.cell(row=cd_row, column=2)
        cell_a = case_data.cell(row=cd_row, column=1)
        print(f"Row {cd_row} (B{cd_row}): {label} = {cell_b.value}")
    
    wb.close()

if __name__ == '__main__':
    script_dir = Path(__file__).parent.absolute()
    template_path = script_dir / 'templates' / 'financial_analysis_template.xlsx'
    check_row_mapping(str(template_path))











