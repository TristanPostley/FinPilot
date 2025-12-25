#!/usr/bin/env python3
"""
Compare template and output Excel files to identify differences.
"""

from openpyxl import load_workbook
from pathlib import Path

def compare_excel_files(template_path, output_path):
    """Compare two Excel files and show differences in key cells."""
    
    print(f"Loading template: {template_path}")
    template_wb = load_workbook(template_path, data_only=False)
    
    print(f"Loading output: {output_path}")
    output_wb = load_workbook(output_path, data_only=False)
    
    # Compare Case Data sheet
    if 'Case Data' in template_wb.sheetnames and 'Case Data' in output_wb.sheetnames:
        print("\n" + "="*80)
        print("COMPARING 'Case Data' SHEET")
        print("="*80)
        
        template_sheet = template_wb['Case Data']
        output_sheet = output_wb['Case Data']
        
        # Key cells to check (metadata rows)
        key_cells_metadata = [
            ('B10', 'Company Name'),
            ('D10', 'Ticker'),
            ('B11', 'Shares Outstanding'),
            ('B12', 'Fiscal Year End'),
        ]
        
        print("\n--- Metadata (Rows 10-12) ---")
        for cell_ref, description in key_cells_metadata:
            template_val = template_sheet[cell_ref].value
            output_val = output_sheet[cell_ref].value
            template_formula = template_sheet[cell_ref].data_type == 'f'
            output_formula = output_sheet[cell_ref].data_type == 'f'
            
            if template_val != output_val or template_formula != output_formula:
                print(f"\n{cell_ref} ({description}):")
                print(f"  Template: {template_val} {'(formula)' if template_formula else ''}")
                print(f"  Output:   {output_val} {'(formula)' if output_formula else ''}")
            else:
                print(f"{cell_ref} ({description}): SAME - {template_val}")
        
        # Income Statement rows (13-24)
        print("\n--- Income Statement Rows (13-24) ---")
        for row_num in range(13, 25):
            cell_ref = f'B{row_num}'
            template_val = template_sheet[cell_ref].value
            output_val = output_sheet[cell_ref].value
            template_formula = template_sheet[cell_ref].data_type == 'f'
            output_formula = output_sheet[cell_ref].data_type == 'f'
            
            if template_val != output_val:
                print(f"\n{cell_ref} (Row {row_num}):")
                print(f"  Template: {template_val} {'(formula)' if template_formula else '(value)'}")
                print(f"  Output:   {output_val} {'(formula)' if output_formula else '(value)'}")
        
        # Balance Sheet rows (25-43)
        print("\n--- Balance Sheet Rows (25-43) ---")
        differences_found = False
        for row_num in range(25, 44):
            cell_ref = f'B{row_num}'
            template_val = template_sheet[cell_ref].value
            output_val = output_sheet[cell_ref].value
            
            if template_val != output_val:
                print(f"{cell_ref} (Row {row_num}): Template={template_val}, Output={output_val}")
                differences_found = True
        
        if not differences_found:
            print("No differences found in Balance Sheet rows (25-43)")
    
    # Compare Financial Statements sheet
    if 'Financial Statements' in template_wb.sheetnames and 'Financial Statements' in output_wb.sheetnames:
        print("\n" + "="*80)
        print("COMPARING 'Financial Statements' SHEET")
        print("="*80)
        
        template_sheet = template_wb['Financial Statements']
        output_sheet = output_wb['Financial Statements']
        
        # Check if Financial Statements sheet has formulas or values
        print("\n--- Checking Financial Statements Structure ---")
        
        # Sample some cells to see if they contain formulas
        sample_cells = ['B5', 'B10', 'B15', 'B20', 'C5', 'C10']
        print("\nSample cells in Financial Statements sheet:")
        for cell_ref in sample_cells:
            try:
                template_cell = template_sheet[cell_ref]
                output_cell = output_sheet[cell_ref]
                
                template_has_formula = template_cell.data_type == 'f'
                output_has_formula = output_cell.data_type == 'f'
                
                if template_has_formula:
                    template_display = f"FORMULA: {template_cell.value}"
                else:
                    template_display = f"VALUE: {template_cell.value}"
                
                if output_has_formula:
                    output_display = f"FORMULA: {output_cell.value}"
                else:
                    output_display = f"VALUE: {output_cell.value}"
                
                print(f"\n{cell_ref}:")
                print(f"  Template: {template_display}")
                print(f"  Output:   {output_display}")
                
                # Check if they reference Case Data
                if template_has_formula and 'Case_Data' in str(template_cell.value):
                    print(f"    -> Template formula references Case Data sheet")
                if output_has_formula and 'Case_Data' in str(output_cell.value):
                    print(f"    -> Output formula references Case Data sheet")
                    
            except Exception as e:
                print(f"{cell_ref}: Error - {e}")
        
        # Check first 30 rows for company name references
        print("\n--- Searching for 'Dollar Tree' in Financial Statements ---")
        found_dollar_tree = False
        for row in range(1, 31):
            for col in range(1, 11):  # Columns A-J
                cell = output_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if 'Dollar Tree' in str(cell.value) or 'DOLLAR TREE' in str(cell.value).upper():
                        print(f"Found 'Dollar Tree' at {cell.coordinate}: {cell.value}")
                        found_dollar_tree = True
        
        if not found_dollar_tree:
            print("No 'Dollar Tree' references found in first 30 rows")
        
        # Check for company name in output
        print("\n--- Searching for company name references in Financial Statements ---")
        case_data_company = output_wb['Case Data']['B10'].value
        if case_data_company:
            print(f"Company name from Case Data B10: {case_data_company}")
            found_company_name = False
            for row in range(1, 31):
                for col in range(1, 11):
                    cell = output_sheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if case_data_company in str(cell.value) or str(case_data_company).upper() in str(cell.value).upper():
                            print(f"Found company name at {cell.coordinate}: {cell.value}")
                            found_company_name = True
                            break
                if found_company_name:
                    break
            if not found_company_name:
                print("Company name not found in Financial Statements sheet")
        
        # Check header row for company name
        print("\n--- Checking Header Rows (1-5) for Company Name ---")
        for row in range(1, 6):
            for col in range(1, 11):
                cell = output_sheet.cell(row=row, column=col)
                if cell.value:
                    cell_str = str(cell.value)
                    if 'Dollar Tree' in cell_str or 'DOLLAR TREE' in cell_str.upper():
                        print(f"Row {row}, Col {col} ({cell.coordinate}): {cell.value}")
                    elif case_data_company and (case_data_company in cell_str or str(case_data_company).upper() in cell_str.upper()):
                        print(f"Row {row}, Col {col} ({cell.coordinate}): {cell.value} (CORRECT COMPANY)")
    
    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Template sheets: {', '.join(template_wb.sheetnames)}")
    print(f"Output sheets: {', '.join(output_wb.sheetnames)}")
    
    template_wb.close()
    output_wb.close()

if __name__ == '__main__':
    # Get paths
    script_dir = Path(__file__).parent.absolute()
    template_path = script_dir / 'templates' / 'financial_analysis_template.xlsx'
    output_path = script_dir / 'Tool_Output' / 'Zoom-Communications-Inc-FY-2025-Financials.xlsx'
    
    compare_excel_files(str(template_path), str(output_path))











